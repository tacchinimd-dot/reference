"""
상품 레퍼런스 보드 — 실시간 협업 서버
실행: python launch.py
"""

import os, json, time, threading
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, send_file
import io
from flask_cors import CORS
from flask_socketio import SocketIO, emit
try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False
import base64 as b64lib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
HTML_FILE  = "product_reference_board.html"
STATE_FILE = os.path.join(BASE_DIR, "board_state.json")
XLSX_PATH  = os.path.join(BASE_DIR, "레퍼런스_백업.xlsx")

# ── 데이터베이스 (PostgreSQL) ────────────────────────────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL", "")
USE_DB = bool(DATABASE_URL)

if USE_DB:
    import psycopg2
    from psycopg2.extras import Json

    def _get_conn():
        return psycopg2.connect(DATABASE_URL, sslmode="require")

    def _init_db():
        with _get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS board_state (
                        id INTEGER PRIMARY KEY DEFAULT 1,
                        state JSONB NOT NULL DEFAULT '{}'::jsonb,
                        updated_at TIMESTAMP DEFAULT NOW()
                    )
                """)
                cur.execute("SELECT COUNT(*) FROM board_state")
                if cur.fetchone()[0] == 0:
                    cur.execute(
                        "INSERT INTO board_state (id, state) VALUES (1, %s)",
                        [Json({"tabs": [{"id": "default", "name": "탭 1", "cards": []}]})]
                    )
            conn.commit()
        print("[DB] PostgreSQL 테이블 초기화 완료")

    _init_db()

# ── 서버 상태 ─────────────────────────────────────────────────────────────────
_state_lock  = threading.Lock()
_user_count  = 0

_DEFAULT_STATE = {"tabs": [{"id": "default", "name": "탭 1", "cards": []}]}

def load_state():
    if USE_DB:
        try:
            with _get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT state FROM board_state WHERE id = 1")
                    row = cur.fetchone()
                    if row and row[0]:
                        return row[0]
        except Exception as e:
            print(f"[DB] load_state 오류: {e}")
        return dict(_DEFAULT_STATE)
    else:
        try:
            if os.path.exists(STATE_FILE):
                with open(STATE_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return dict(_DEFAULT_STATE)

def save_state(state):
    with _state_lock:
        if USE_DB:
            try:
                with _get_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE board_state SET state = %s, updated_at = NOW() WHERE id = 1",
                            [Json(state)]
                        )
                    conn.commit()
            except Exception as e:
                print(f"[DB] save_state 오류: {e}")
        else:
            with open(STATE_FILE, "w", encoding="utf-8") as f:
                json.dump(state, f, ensure_ascii=False, indent=2)

board_state = load_state()

# ── HTTP Routes ───────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(BASE_DIR, HTML_FILE)

@app.route("/ping")
def ping():
    return jsonify({"ok": True})

@app.route("/backup", methods=["POST"])
def backup():
    try:
        save_to_xlsx(request.get_json())
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/export", methods=["GET"])
def export_xlsx():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "레퍼런스"

        ws.column_dimensions["A"].width = 20   # 이미지
        ws.column_dimensions["B"].width = 10   # 라벨
        ws.column_dimensions["C"].width = 40   # 값

        thin       = Side(style="thin", color="E5E7EB")
        card_border = Border(top=thin, bottom=thin, left=thin, right=thin)
        label_fill  = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        tab_fill    = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")

        CARD_ROWS    = 4
        ROW_HEIGHT   = 28
        IMG_W, IMG_H = 120, 100

        row = 1
        with _state_lock:
            tabs = board_state.get("tabs", [])

        for tab in tabs:
            cards = [c for c in tab.get("cards", []) if (c.get("type") or "card") == "card"]
            if not cards:
                continue

            # 탭 헤더
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            hdr = ws.cell(row=row, column=1, value=f"  {tab.get('name','')}")
            hdr.font      = Font(bold=True, color="FFFFFF", size=12)
            hdr.fill      = tab_fill
            hdr.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 28
            row += 1

            for card in cards:
                start = row

                # 행 높이 설정
                for r in range(CARD_ROWS):
                    ws.row_dimensions[start + r].height = ROW_HEIGHT

                # 이미지 셀 (세로 병합)
                ws.merge_cells(start_row=start, start_column=1,
                               end_row=start + CARD_ROWS - 1, end_column=1)
                img_cell = ws.cell(row=start, column=1)
                img_cell.alignment = Alignment(horizontal="center", vertical="center")
                img_cell.border    = card_border

                # 이미지 임베딩
                image_val = card.get("image", "")
                if image_val and image_val.startswith("data:image"):
                    try:
                        raw = b64lib.b64decode(image_val.split(",", 1)[1])
                        pil = PILImage.open(io.BytesIO(raw))
                        if pil.mode not in ("RGB", "RGBA"):
                            pil = pil.convert("RGBA")
                        out = io.BytesIO()
                        pil.save(out, format="PNG")
                        out.seek(0)
                        xl_img        = XLImage(out)
                        xl_img.width  = IMG_W
                        xl_img.height = IMG_H
                        ws.add_image(xl_img, f"A{start}")
                    except Exception:
                        img_cell.value = "🖼️"
                else:
                    img_cell.value = "—"

                # 상세 정보 (4행)
                details = [
                    ("상품명", card.get("name", ""),     False),
                    ("가격",   card.get("price", ""),    False),
                    ("소재",   card.get("material", ""), False),
                    ("URL",    card.get("url", ""),      True),
                ]
                for i, (label, value, is_url) in enumerate(details):
                    r = start + i

                    lbl = ws.cell(row=r, column=2, value=label)
                    lbl.font      = Font(bold=True, size=10, color="6B7280")
                    lbl.fill      = label_fill
                    lbl.alignment = Alignment(horizontal="center", vertical="center")
                    lbl.border    = card_border

                    val = ws.cell(row=r, column=3, value=value)
                    val.alignment = Alignment(vertical="center", wrap_text=True)
                    val.border    = card_border
                    if is_url and value:
                        val.hyperlink = value
                        val.font = Font(size=11, color="6366F1", underline="single")
                    else:
                        val.font = Font(size=11)

                row += CARD_ROWS

                # 카드 사이 구분선
                ws.row_dimensions[row].height = 6
                row += 1

            # 탭 사이 여백
            ws.row_dimensions[row].height = 14
            row += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"레퍼런스_카드_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/scrape")
def scrape():
    url = request.args.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL이 필요합니다"}), 400

    on_render = bool(os.environ.get("RENDER"))
    result = None

    # 1차: Playwright 시도 (Render 환경 또는 미설치 시 스킵)
    if not on_render and HAS_PLAYWRIGHT:
        try:
            result = scrape_product(url)
        except Exception:
            pass

    # 2차: 결과가 없거나 이미지+상품명 모두 비어 있으면 HTTP fallback
    if not result or (not result.get("image") and not result.get("name")):
        try:
            result = scrape_fallback(url)
        except Exception:
            pass

    # 3차: 두 방법 모두 부분 결과면 병합 (Playwright + fallback)
    if not on_render and result and (not result.get("image") or not result.get("name")):
        try:
            fb = scrape_fallback(url)
            for k in ("image", "name", "price", "material"):
                if not result.get(k) and fb.get(k):
                    result[k] = fb[k]
        except Exception:
            pass

    if not result:
        return jsonify({"error": "상품 정보를 가져올 수 없습니다"}), 500
    return jsonify(result)

# ── Socket.IO Events ──────────────────────────────────────────────────────────
@socketio.on("connect")
def on_connect():
    global _user_count
    _user_count += 1
    # 연결한 클라이언트에게 전체 상태 전송
    emit("state_sync", {"state": board_state, "userCount": _user_count})
    # 나머지 클라이언트에게 접속자 수 업데이트
    emit("user_count", {"count": _user_count}, broadcast=True, include_self=False)
    print(f"[+] 클라이언트 연결 ({_user_count}명 접속 중): {request.sid}")

@socketio.on("disconnect")
def on_disconnect():
    global _user_count
    _user_count = max(0, _user_count - 1)
    emit("user_count", {"count": _user_count}, broadcast=True)
    print(f"[-] 클라이언트 연결 해제 ({_user_count}명 접속 중): {request.sid}")

@socketio.on("state_upload")
def on_state_upload(data):
    """클라이언트가 로컬 데이터를 서버에 업로드 (서버 상태가 비어있을 때)"""
    global board_state
    if not any(t.get("cards") for t in board_state.get("tabs", [])):
        board_state = data
        save_state(board_state)
        emit("state_sync", {"state": board_state, "userCount": _user_count}, broadcast=True)

@socketio.on("state_replace")
def on_state_replace(data):
    """클라이언트가 강제로 전체 상태를 교체 (Undo 등 상태 전체 변경 시)"""
    global board_state
    board_state = data
    save_state(board_state)
    emit("state_sync", {"state": board_state, "userCount": _user_count}, broadcast=True, include_self=False)

@socketio.on("item_add")
def on_item_add(data):
    # data: {tabId, item}
    tab = _find_tab(data["tabId"])
    if tab:
        tab["cards"].append(data["item"])
        save_state(board_state)
        emit("item_add", data, broadcast=True, include_self=False)

@socketio.on("item_update")
def on_item_update(data):
    # data: {tabId, itemId, patch}
    tab = _find_tab(data["tabId"])
    if tab:
        item = next((i for i in tab["cards"] if str(i["id"]) == str(data["itemId"])), None)
        if item:
            item.update(data["patch"])
            save_state(board_state)
            emit("item_update", data, broadcast=True, include_self=False)

@socketio.on("item_delete")
def on_item_delete(data):
    # data: {tabId, itemId}
    tab = _find_tab(data["tabId"])
    if tab:
        tab["cards"] = [i for i in tab["cards"] if str(i["id"]) != str(data["itemId"])]
        save_state(board_state)
        emit("item_delete", data, broadcast=True, include_self=False)

@socketio.on("tab_add")
def on_tab_add(data):
    board_state["tabs"].append(data["tab"])
    save_state(board_state)
    emit("tab_add", data, broadcast=True, include_self=False)

@socketio.on("tab_delete")
def on_tab_delete(data):
    board_state["tabs"] = [t for t in board_state["tabs"] if t["id"] != data["tabId"]]
    if not board_state["tabs"]:
        board_state["tabs"] = [{"id": "default", "name": "탭 1", "cards": []}]
    save_state(board_state)
    emit("tab_delete", data, broadcast=True, include_self=False)

@socketio.on("tab_rename")
def on_tab_rename(data):
    tab = _find_tab(data["tabId"])
    if tab:
        tab["name"] = data["name"]
        save_state(board_state)
        emit("tab_rename", data, broadcast=True, include_self=False)

@socketio.on("tab_clear")
def on_tab_clear(data):
    tab = _find_tab(data["tabId"])
    if tab:
        tab["cards"] = []
        save_state(board_state)
        emit("tab_clear", data, broadcast=True, include_self=False)

def _find_tab(tab_id):
    return next((t for t in board_state["tabs"] if t["id"] == tab_id), None)

# ── Scraper ───────────────────────────────────────────────────────────────────
def scrape_product(url: str) -> dict:
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ],
        )
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
            locale="ko-KR",
            extra_http_headers={
                "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
                "Sec-CH-UA": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
                "Sec-CH-UA-Mobile": "?0",
                "Sec-CH-UA-Platform": '"Windows"',
            },
        )
        page = ctx.new_page()
        # navigator.webdriver 숨기기
        page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
            Object.defineProperty(navigator, 'languages', { get: () => ['ko-KR', 'ko', 'en-US', 'en'] });
            window.chrome = { runtime: {} };
        """)
        page.goto(url, wait_until="domcontentloaded", timeout=25000)
        try:
            page.wait_for_load_state("networkidle", timeout=7000)
        except Exception:
            pass
        # SPA(React/Next.js) 렌더링 완료 대기
        try:
            page.wait_for_selector('[class*="price"],[class*="Price"],[itemprop="price"]', timeout=5000)
        except Exception:
            page.wait_for_timeout(2000)
        raw = page.evaluate(_EXTRACT_JS)
        browser.close()
    return _post_process(raw, url)

def scrape_fallback(url: str) -> dict:
    import urllib.request, re
    from bs4 import BeautifulSoup
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Sec-CH-UA": '"Google Chrome";v="131", "Chromium";v="131"',
    })
    html = urllib.request.urlopen(req, timeout=15).read().decode("utf-8", errors="ignore")
    soup = BeautifulSoup(html, "html.parser")

    result = {"name": "", "price": "", "material": "", "image": ""}

    # ── 1) ld+json Product 스키마 파싱 ──
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            d = json.loads(script.string or "")
            items = [d] + (d.get("@graph") or [])
            for item in items:
                if item.get("@type") != "Product":
                    continue
                result["name"] = result["name"] or item.get("name", "")
                imgs = item.get("image", "")
                if isinstance(imgs, list):
                    result["image"] = result["image"] or (imgs[0] if imgs else "")
                elif isinstance(imgs, dict):
                    result["image"] = result["image"] or imgs.get("url", "")
                else:
                    result["image"] = result["image"] or str(imgs)
                offers = item.get("offers", {})
                if isinstance(offers, list):
                    offers = offers[0] if offers else {}
                if not result["price"] and offers.get("price"):
                    cur = offers.get("priceCurrency", "")
                    result["price"] = f"{cur} {offers['price']}".strip()
                if not result["material"] and item.get("material"):
                    mat = item["material"]
                    result["material"] = mat if isinstance(mat, str) else (mat.get("name", ""))
                if not result["material"] and item.get("description"):
                    result["material"] = _extract_material(item["description"])
        except Exception:
            pass

    # ── 2) __NEXT_DATA__ (Next.js SPA) ──
    next_script = soup.find("script", id="__NEXT_DATA__")
    if next_script and next_script.string:
        try:
            nd = json.loads(next_script.string)
            pp = (nd.get("props") or {}).get("pageProps") or {}
            prod = (pp.get("product") or pp.get("productDetail") or pp.get("item")
                    or pp.get("goodsDetail") or (pp.get("initialState") or {}).get("product")
                    or (pp.get("initialState") or {}).get("goods"))
            if not prod:
                queries = ((pp.get("dehydratedState") or {}).get("queries") or [{}])
                qdata = (queries[0].get("state") or {}).get("data") or {}
                prod = qdata.get("data") or qdata
            if prod and isinstance(prod, dict):
                result["name"] = result["name"] or prod.get("name") or prod.get("title") or prod.get("goodsName") or prod.get("productName") or ""
                raw_price = prod.get("price") or prod.get("salePrice") or prod.get("sellingPrice") or prod.get("retailPrice") or ""
                if not result["price"] and raw_price:
                    result["price"] = str(raw_price)
                result["material"] = result["material"] or prod.get("material") or prod.get("fabric") or prod.get("composition") or ""
                img_val = prod.get("image") or prod.get("imageUrl") or prod.get("thumbnail") or prod.get("listImageUrl") or prod.get("mainImageUrl") or ""
                if isinstance(img_val, list):
                    img_val = img_val[0] if img_val else ""
                if isinstance(img_val, dict):
                    img_val = img_val.get("url") or img_val.get("src") or img_val.get("imageUrl") or ""
                result["image"] = result["image"] or str(img_val)
        except Exception:
            pass

    # ── 3) Open Graph / meta 태그 ──
    def find_meta(prop):
        for prefix in ("", "og:", "twitter:"):
            tag = soup.find("meta", attrs={"property": f"{prefix}{prop}"}) or soup.find("meta", attrs={"name": f"{prefix}{prop}"})
            if tag and tag.get("content"):
                return tag["content"].strip()
        return ""

    result["name"]  = result["name"]  or find_meta("title") or (soup.find("h1").get_text(strip=True) if soup.find("h1") else "")
    result["image"] = result["image"] or find_meta("image")
    if not result["price"]:
        result["price"] = find_meta("price:amount") or find_meta("price")
        if result["price"] and not re.search(r'[\u20a9$€£¥]', result["price"]) and not re.match(r'^[A-Z]{3}\s', result["price"]):
            cur = find_meta("price:currency")
            if cur:
                result["price"] = f"{cur} {result['price']}"

    # ── 4) HTML 구조에서 가격 추출 ──
    if not result["price"]:
        price_sels = [
            {"attrs": {"itemprop": "price"}},
            {"class_": re.compile(r"sale.*price|price.*sale|current.*price|final.*price|discount.*price", re.I)},
            {"attrs": {"data-testid": re.compile(r"price", re.I)}},
        ]
        for sel in price_sels:
            tag = soup.find(**sel)
            if tag:
                txt = tag.get_text(strip=True)
                if re.search(r'\d', txt) and len(txt) < 35:
                    result["price"] = txt
                    break

    # ── 5) 소재 추출 ──
    if not result["material"]:
        mat_tag = soup.find(attrs={"itemprop": "material"}) or soup.find(class_=re.compile(r"material|fabric|composition", re.I))
        if mat_tag:
            result["material"] = mat_tag.get_text(strip=True)[:120]
    if not result["material"]:
        result["material"] = _extract_material(soup.get_text())

    for k in result:
        result[k] = re.sub(r'\s+', ' ', (result[k] or "")).strip()
    return _post_process(result, url)


def _extract_material(text: str) -> str:
    """텍스트에서 소재 정보 추출"""
    if not text:
        return ""
    import re
    m1 = re.search(r'\d+%\s*[A-Za-z\uac00-\ud7a3]+(?:\s*[,/·]\s*\d+%\s*[A-Za-z\uac00-\ud7a3]+)+', text)
    if m1:
        return m1.group(0).strip()[:120]
    m2 = re.search(r'(?:소재|원단|재질|fabric|material|composition)\s*[:\s]+([^\n.<]{4,100})', text, re.I)
    if m2:
        val = m2.group(1).strip()
        if re.match(r'^(JSON|HTML|CSS|script|function|var |const |let |http)', val, re.I):
            return ""
        return val[:120]
    return ""

_EXTRACT_JS = r"""
() => {
  const result = { name: '', price: '', material: '', image: '' };

  // ── Next.js / __NEXT_DATA__ (SPA 우선 추출) ──
  try {
    const nd = document.getElementById('__NEXT_DATA__');
    if (nd) {
      const parsed = JSON.parse(nd.textContent);
      const pp = parsed.props?.pageProps || {};
      // 29cm, 무신사 등 공통 경로 탐색
      const prod = pp.product || pp.productDetail || pp.item || pp.goodsDetail ||
                   pp.initialState?.product || pp.initialState?.goods ||
                   pp.dehydratedState?.queries?.[0]?.state?.data ||
                   pp.dehydratedState?.queries?.[0]?.state?.data?.data;
      if (prod) {
        result.name  = result.name  || prod.name || prod.title || prod.goodsName || prod.productName || '';
        const rawPrice = prod.price || prod.salePrice || prod.sellingPrice || prod.retailPrice || prod.finalPrice || '';
        if (!result.price && rawPrice) result.price = String(rawPrice).replace(/[^0-9,]/g,'') ? String(rawPrice) : '';
        result.material = result.material || prod.material || prod.fabric || prod.composition || '';
        const imgs = prod.image || prod.imageUrl || prod.thumbnail ||
                     prod.listImageUrl || prod.mainImageUrl ||
                     (prod.images && (Array.isArray(prod.images) ? prod.images[0] : prod.images)) || '';
        result.image = result.image || (typeof imgs === 'string' ? imgs : (imgs?.url || imgs?.src || imgs?.imageUrl || ''));
      }
    }
  } catch(e) {}

  // ── ld+json Product 스키마 ──
  for (const s of document.querySelectorAll('script[type="application/ld+json"]')) {
    try {
      const d = JSON.parse(s.textContent);
      const items = [d, ...(d['@graph'] || [])];
      for (const item of items) {
        if (item['@type'] !== 'Product') continue;
        result.name  = result.name  || item.name  || '';
        const imgs = item.image;
        result.image = result.image || (Array.isArray(imgs) ? imgs[0] : (typeof imgs === 'object' ? imgs.url : imgs)) || '';
        const offers = [].concat(item.offers || [])[0] || {};
        if (!result.price && offers.price)
          result.price = ((offers.priceCurrency || '') + ' ' + offers.price).trim();
        if (!result.material && item.description) result.material = extractMaterial(item.description);
        if (!result.material && item.material) result.material = typeof item.material === 'string' ? item.material : (item.material.name || '');
      }
    } catch(e) {}
  }

  const m = p => (document.querySelector(`meta[property="${p}"],meta[name="${p}"]`) || {}).content || '';
  result.name  = result.name  || m('og:title')  || m('twitter:title')  || (document.querySelector('h1') || {}).textContent || '';
  result.image = result.image || m('og:image')  || m('twitter:image')  || '';
  result.price = result.price || m('product:price:amount') || m('og:price:amount') || '';
  if (result.price && !/[\u20a9$\u20ac\xa3\xa5]/.test(result.price) && !/^[A-Z]{3}\s/.test(result.price)) {
    const cur = m('product:price:currency') || m('og:price:currency');
    if (cur) result.price = cur + ' ' + result.price;
  }

  if (!result.price) {
    const sels = [
      '[itemprop="price"]',
      '[class*="salePrice" i]','[class*="sale_price" i]','[class*="FinalPrice" i]',
      '[class*="sale"][class*="price" i]','[class*="current"][class*="price" i]',
      '[data-testid*="price" i]','.price--sale','.price__current',
      '[class*="Price"][class*="discount" i]','[class*="discountPrice" i]',
    ];
    for (const sel of sels) {
      try { const el = document.querySelector(sel); if (el) { const t = el.textContent.trim(); if (/\d/.test(t) && t.length < 35) { result.price = t; break; } } } catch(e) {}
    }
  }

  if (!result.material) {
    const sels = ['[itemprop="material"]','[class*="material" i]','[class*="fabric" i]','[class*="composition" i]'];
    for (const sel of sels) { try { const el = document.querySelector(sel); if (el) { result.material = el.textContent.trim().slice(0,120); break; } } catch(e) {} }
  }
  if (!result.material) result.material = extractMaterial(document.body.textContent);

  if (!result.image) {
    const sels = ['[class*="product"][class*="image" i] img','[class*="gallery" i] img','[class*="hero" i] img'];
    for (const sel of sels) {
      try { const imgs = [...document.querySelectorAll(sel)].filter(i => i.naturalWidth > 200 && i.src && !i.src.includes('data:')); if (imgs.length) { result.image = imgs[0].src; break; } } catch(e) {}
    }
  }

  function extractMaterial(text) {
    if (!text) return '';
    const m1 = text.match(/\d+%\s*[A-Za-z\uac00-\ud7a3]+(?:\s*[,\/\u00b7]\s*\d+%\s*[A-Za-z\uac00-\ud7a3]+)+/);
    if (m1) return m1[0].trim().slice(0, 120);
    const m2 = text.match(/(?:\uc18c\uc7ac|\uc6d0\ub2e8|\uc7ac\uc9c8|fabric|material|composition)\s*[:\s]+([^\n.<]{4,100})/i);
    if (m2) {
      const val = m2[1].trim();
      // 잘못된 매칭 필터링 (JSON, script 등 기술 용어)
      if (/^(JSON|HTML|CSS|script|function|var |const |let |http)/i.test(val)) return '';
      return val.slice(0, 120);
    }
    return '';
  }

  for (const k of Object.keys(result)) result[k] = (result[k] || '').replace(/\s+/g, ' ').trim();
  return result;
}
"""

def _post_process(raw: dict, src_url: str) -> dict:
    img = raw.get("image", "")
    if img:
        if img.startswith("//"): img = "https:" + img
        elif not img.startswith("http"):
            from urllib.parse import urljoin
            img = urljoin(src_url, img)
    raw["image"] = img
    return raw

# ── xlsx 백업 ─────────────────────────────────────────────────────────────────
HEADERS   = ["추가일시", "탭", "상품명", "가격", "소재", "URL"]

def _init_xlsx():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "레퍼런스"
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="6366f1")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for w, c in zip([18,14,36,14,32,60], "ABCDEF"): ws.column_dimensions[c].width = w
    ws.row_dimensions[1].height = 22
    wb.save(XLSX_PATH)

def save_to_xlsx(data: dict):
    if not os.path.exists(XLSX_PATH): _init_xlsx()
    wb = openpyxl.load_workbook(XLSX_PATH); ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), data.get("tab",""), data.get("name",""), data.get("price",""), data.get("material",""), data.get("url","")])
    url_val = data.get("url", "")
    if url_val:
        cell = ws.cell(row=ws.max_row, column=6)
        cell.hyperlink = url_val; cell.font = Font(color="6366f1", underline="single")
    wb.save(XLSX_PATH)

# ── Entry ─────────────────────────────────────────────────────────────────────
def start(open_browser=False):
    port = int(os.environ.get("PORT", 5000))
    on_render = bool(os.environ.get("RENDER"))
    if open_browser and not on_render:
        def _open():
            import subprocess
            subprocess.Popen(["cmd", "/c", "start", f"http://localhost:{port}"], shell=False)
        threading.Timer(1.5, _open).start()
    print(f"  [저장소] {'PostgreSQL (외부 DB)' if USE_DB else 'board_state.json (로컬 파일)'}")
    if not on_render:
        import socket as _sock
        host_ip = _sock.gethostbyname(_sock.gethostname())
        print("=" * 54)
        print("  상품 레퍼런스 보드  (실시간 협업)")
        print(f"  내 PC:    http://localhost:{port}")
        print(f"  같은 네트워크 팀원:  http://{host_ip}:{port}")
        print(f"  저장소:   {'PostgreSQL' if USE_DB else '로컬 파일'}")
        print("  종료하려면 Ctrl+C 또는 창을 닫으세요")
        print("=" * 54)
    socketio.run(app, host="0.0.0.0", port=port, debug=False, allow_unsafe_werkzeug=True)

if __name__ == "__main__":
    start(open_browser=True)
